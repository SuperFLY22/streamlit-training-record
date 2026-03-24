import os
import io
import base64
import zipfile
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as xlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.cell import coordinate_to_tuple
from PIL import Image as PILImage
from PIL import ImageChops

TEMPLATE_PATH = "public/template.xlsx"

def base64_to_image(b64_str):
    if "," in b64_str:
        b64_str = b64_str.split(",")[1]
    img_data = base64.b64decode(b64_str)
    return PILImage.open(io.BytesIO(img_data))


def process_signature(b64_str, target_w, target_h, margin_px=4):
    """
    1. 서명 캔버스의 흰색 여백을 잘라내 실제 서명 획만 추출
    2. (target - margin*2) 크기에 맞춰 비율 유지 축소
    3. 결과물을 흰색 캔버스 중앙에 배치
    4. margin_px 만큼 셀 안쪽에서 띄워 배치(openpyxl OneCellAnchor offset)
       → 셀 테두리가 이미지에 가려지지 않음
    """
    img_pil = base64_to_image(b64_str)

    # 1. 흰 여백 트리밍 (실서명만 추출)
    img_rgb = img_pil.convert("RGB")
    bg = PILImage.new("RGB", img_rgb.size, (255, 255, 255))
    diff = ImageChops.difference(img_rgb, bg)
    bbox = diff.getbbox()
    if bbox:
        img_pil = img_pil.crop(bbox)

    # 2. 이미지가 배치될 실제 영역 = 셀 - 여백
    inner_w = max(target_w - margin_px * 2, 1)
    inner_h = max(target_h - margin_px * 2, 1)

    # 비율 유지하여 inner 사이즈에 맞게 축소
    img_pil.thumbnail((inner_w, inner_h), PILImage.Resampling.LANCZOS)

    # 3. inner 사이즈의 흰색 캔버스를 만들고 가운데 붙이기
    canvas = PILImage.new("RGBA", (inner_w, inner_h), (255, 255, 255, 255))
    paste_x = (inner_w - img_pil.width) // 2
    paste_y = (inner_h - img_pil.height) // 2
    canvas.paste(img_pil, (max(0, paste_x), max(0, paste_y)))

    img_byte_arr = io.BytesIO()
    canvas.save(img_byte_arr, format='PNG')
    img_byte_arr.seek(0)

    xl_img = xlImage(img_byte_arr)
    xl_img.width  = inner_w
    xl_img.height = inner_h
    return xl_img


def add_image_inset(ws, xl_img, cell_ref, margin_px=4):
    """
    OpenpyXL OneCellAnchor를 직접 설정하여
    이미지를 셀의 (margin_px, margin_px) 위치부터 시작하도록 배치.
    → 좌측·상단 셀 테두리가 이미지에 가려지지 않음
    """
    row_idx, col_idx = coordinate_to_tuple(cell_ref)  # 1-indexed 반환

    # 1 pixel = 9525 EMU (96 DPI 기준)
    offset_emu = int(margin_px * 9525)
    img_w_emu  = int(xl_img.width  * 9525)
    img_h_emu  = int(xl_img.height * 9525)

    marker = AnchorMarker(
        col    = col_idx - 1,   # 0-indexed
        colOff = offset_emu,
        row    = row_idx - 1,   # 0-indexed
        rowOff = offset_emu,
    )
    size   = XDRPositiveSize2D(cx=img_w_emu, cy=img_h_emu)
    anchor = OneCellAnchor(_from=marker, ext=size)
    xl_img.anchor = anchor
    ws.add_image(xl_img)


def generate_excel_or_zip(subject_name, subject_content, course_time, lecture_date,
                          submitted_date, instructor_info, trainees):
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template file not found at {TEMPLATE_PATH}")

    # 15명 단위로 분할
    chunk_size = 15
    trainee_chunks = [trainees[i:i + chunk_size]
                      for i in range(0, max(1, len(trainees)), chunk_size)]

    files = []

    for idx, chunk in enumerate(trainee_chunks):
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        ws['B3'] = subject_name
        ws['B4'] = subject_content
        ws['F5'] = course_time
        ws['B5'] = lecture_date if lecture_date else ""
        ws['C27'] = submitted_date if submitted_date else ""
        ws['I5']  = instructor_info.get("location", "")
        ws['B7']  = instructor_info.get("company", "")
        ws['F7']  = instructor_info.get("instructor_id", "")
        ws['I7']  = instructor_info.get("name", "")
        ws['F8']  = instructor_info.get("course_name", "")

        # 강사 서명 (I+J 병합, 약 6cm × 행높이)
        instructor_sig = instructor_info.get("signature")
        if instructor_sig:
            xl_img = process_signature(instructor_sig, target_w=227, target_h=53)
            add_image_inset(ws, xl_img, 'I27')

        # 훈련생 서명 (H+I 병합, 4cm × 0.85cm)
        row_start = 10
        for i, trainee in enumerate(chunk):
            row = row_start + i
            ws[f'A{row}'] = trainee.get("team", "")
            ws[f'C{row}'] = trainee.get("employee_id", "")
            ws[f'E{row}'] = trainee.get("name", "")

            trainee_sig = trainee.get("signature")
            if trainee_sig:
                t_xl_img = process_signature(trainee_sig, target_w=151, target_h=32)
                add_image_inset(ws, t_xl_img, f'H{row}')

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        label = f"Report_Part{idx+1}.xlsx" if len(trainee_chunks) > 1 else "Report.xlsx"
        files.append((label, output.getvalue()))

    if len(files) == 1:
        return files[0][1], "xlsx"
    else:
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for fname, fbytes in files:
                zf.writestr(fname, fbytes)
        zip_buf.seek(0)
        return zip_buf.getvalue(), "zip"
